import os, base64, json, re
import pypdf
from datetime import datetime, timedelta
from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import anthropic
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__)
CORS(app)

ANTHROPIC_API_KEY = os.environ.get('ANTHROPIC_API_KEY', '')

# ── BUDGET LINE DEFINITIONS ───────────────────────────────────────────────────

# ── PHASE NAME NORMALIZATION ──────────────────────────────────────────────────
PHASE_ALIASES = {
    'CASTING': [
        'CASTING', 'CASTING PERIOD', 'TALENT SEARCH', 'AUDITIONS',
        'CALLBACKS', 'CASTING ROUNDS', 'CASTING SESSIONS',
    ],
    'PREP': [
        'PREP', 'PRE-PRODUCTION', 'PRE PRODUCTION', 'PRE-PROD', 'PRE PROD',
        'FIELD PREP', 'STUDIO PREP', 'FIELD PRANKS PREP', 'PREP FIELD PRANKS',
        'OFFICE OPEN', 'START STAFF', 'STAFF START', 'OFFICE START',
        'SCOUT', 'TECH SCOUT', 'LOCATION SCOUT', 'SCOUTS',
        'READ THROUGH', 'TABLE READ',
    ],
    'PRODUCTION': [
        'LOAD', 'LOAD IN', 'LOAD-IN', 'LOADIN', 'LOAD OUT', 'LOAD-OUT',
        'LOADOUT', 'ESU', 'EQUIPMENT SET UP', 'EQUIPMENT SETUP',
        'SET UP', 'SETUP', 'TECH SET UP', 'TECH SETUP', 'RIG',
        'SHOOT', 'SHOOT DAY', 'SHOOTING', 'PRODUCTION',
        'SHOOT PITCHES', 'SHOOT PRANKS', 'SHOOT EPISODES',
        'SHOOT PRANK', 'SHOOT PITCH', 'SHOOT EPISODE',
        'FIELD SHOOT', 'STUDIO SHOOT', 'FIELD PRANKS', 'FIELD PRANK',
        'PRANK STAGE', 'PITCH STAGE', 'STUDIO STAGE',
        'TAPING', 'TAPE', 'FILMING', 'FILM DAY',
        'PRINCIPAL PHOTOGRAPHY', 'RECORD', 'RECORDING', 'RECORDING DAY',
        'REHEARSE', 'REHEARSAL', 'REHEARSALS', 'TECH REHEARSE',
        'TECH REHEARSAL', 'DRY RUN', 'RUN THROUGH', 'RUN-THROUGH',
        'CAMERA REHEARSAL', 'DRESS REHEARSAL', 'BLOCKING',
        'VTR', 'PLAYBACK', 'PLAYBACK DAY',
        'TRAVEL', 'TRAVEL DAY', 'TRAVEL DAYS',
        'DARK', 'DARK DAY', 'DARK WEEK', 'DOWN DAY',
        'STRIKE', 'STRIKE DAY', 'WRAP DAY', 'SET STRIKE',
        'HIATUS', 'BREAK', 'HOLIDAY BREAK', 'CHRISTMAS', 'THANKSGIVING',
        'FIELD PRANKS X', 'FIELD PRANKS WEEK', 'BTS', 'BEHIND THE SCENES',
        'LOCATION SHOOT', 'ON LOCATION', 'REMOTE SHOOT',
    ],
    'POST': [
        'POST', 'POST PRODUCTION', 'POST-PRODUCTION', 'POST-PROD', 'POST PROD',
        'EDIT', 'EDITING', 'EDITORIAL', 'OFFLINE', 'OFFLINE EDIT',
        'ONLINE', 'ONLINE EDIT', 'COLOR', 'COLOR CORRECT', 'COLOR CORRECTION',
        'MIX', 'AUDIO MIX', 'SOUND MIX', 'FINISHING', 'POST FINISHING',
        'DELIVERY', 'MASTER DELIVERY', 'DELIVERIES',
        'QC', 'QUALITY CONTROL', 'QUALITY CHECK',
        'CLOSED CAPTIONING', 'CAPTIONING', 'CC',
        'VFX', 'VISUAL EFFECTS', 'GRAPHICS',
        'WRAP', 'POST WRAP', 'PRODUCTION WRAP',
        'TRANSCRIPTION', 'TRANSCRIPTIONS',
    ],
}

PRODUCTION_CONTAINS = [
    'SHOOT', 'LOAD', 'ESU', 'FIELD PRANK', 'PRANK STAGE', 'PITCH STAGE',
    'REHEARS', 'DRY RUN', 'RUN THROUGH', 'STRIKE', 'HIATUS', 'DARK',
    'TRAVEL', 'TAPING', 'FILMING', 'VTR', 'PLAYBACK', 'BTS',
]
PREP_CONTAINS     = ['PREP', 'SCOUT', 'PRE-PROD', 'PRE PROD']
CASTING_CONTAINS  = ['CASTING', 'AUDITION', 'CALLBACK']
POST_CONTAINS     = ['POST', 'EDIT', 'COLOR', 'MIX', 'DELIVER', 'QC', 'CAPTIO']

def canonical_phase(name):
    upper = name.upper().strip()
    for canon, aliases in PHASE_ALIASES.items():
        if upper in aliases:
            return canon
    if any(kw in upper for kw in PRODUCTION_CONTAINS): return 'PRODUCTION'
    if any(kw in upper for kw in CASTING_CONTAINS):    return 'CASTING'
    if any(kw in upper for kw in POST_CONTAINS):       return 'POST'
    if any(kw in upper for kw in PREP_CONTAINS):       return 'PREP'
    return None

# ── CALENDAR HELPERS ──────────────────────────────────────────────────────────

def get_monday(d):
    return d - timedelta(days=d.weekday())

def wk_idx(d, wk1):
    return round((get_monday(d) - wk1).days / 7)

def parse_date(s):
    for fmt in ('%Y-%m-%d','%m/%d/%Y','%m/%d/%y'):
        try:
            return datetime.strptime(s.strip(), fmt)
        except ValueError:
            continue
    return None

# ── SPREAD ENGINE ─────────────────────────────────────────────────────────────

def build_phase_ranges(phases, wk1, nw):
    ranges = {}
    for ph in phases:
        canon = canonical_phase(ph['name'])
        if not canon:
            continue
        s = parse_date(ph['start'])
        e = parse_date(ph['end'])
        if not s or not e:
            continue
        si = max(0, wk_idx(s, wk1))
        ei = min(nw-1, wk_idx(e, wk1))
        if canon in ranges:
            ranges[canon] = (min(ranges[canon][0], si), max(ranges[canon][1], ei))
        else:
            ranges[canon] = (si, ei)
    return ranges

def spread(spread_type, amt, nw, pr):
    w = [0.0] * nw
    if not amt:
        return w

    def fill(s, e, a):
        s = max(0, s); e = min(nw-1, e)
        if s > e: return
        pw = a / (e - s + 1)
        for i in range(s, e+1):
            w[i] += pw

    cs, ce   = pr.get('CASTING',    (0, 0))
    ps, pe   = pr.get('PREP',       (0, 0))
    ss, se   = pr.get('PRODUCTION', (0, 0))
    pos, poe = pr.get('POST',       (0, nw-1))

    if   spread_type == 'full':              fill(0, nw-1, amt)
    elif spread_type == 'casting':           fill(cs, ce, amt)
    elif spread_type == 'prep_only':         fill(ps, pe, amt)
    elif spread_type == 'prep_shoot':        fill(ps, se, amt)
    elif spread_type == 'prep_shoot_plus4':  fill(ps, se+4, amt)
    elif spread_type == 'talent_fees':       fill(max(0, ss-3), se, amt)
    elif spread_type == 'shoot_only':        fill(ss, se, amt)
    elif spread_type == 'shoot_plus2':       fill(max(0, ss-2), se, amt)
    elif spread_type == 'shoot_plus4':       fill(max(0, ss-4), se, amt)
    elif spread_type == 'post':              fill(pos, poe, amt)
    elif spread_type == 'post_plus2':        fill(max(0, pos-2), poe, amt)
    elif spread_type == 'graphics':
        fill(max(0, ss-4), pos-1, amt/2)
        fill(pos, poe, amt/2)
    elif spread_type == 'week1':             w[0] += amt
    else:                                    fill(0, nw-1, amt)
    return w

# ── PAYMENT ENGINE ────────────────────────────────────────────────────────────

def compute_payments(wk_totals, nw, pr, phases):
    cs, ce   = pr.get('CASTING',    (0, 0))
    ps, pe   = pr.get('PREP',       (0, 0))
    ss, se   = pr.get('PRODUCTION', (0, 0))
    pos, poe = pr.get('POST',       (0, nw-1))
    mid_prep = (ps + pe) // 2
    mid_post = (pos + poe) // 2

    ms = []; labels = []
    ms.append(cs if 'CASTING' in pr else min(v[0] for v in pr.values()))
    labels.append('Start of Casting' if 'CASTING' in pr else 'Start of Production')
    if 'PREP' in pr:
        ms.append(mid_prep); labels.append('Mid Prep')
    else:
        ms.append(min(v[0] for v in pr.values()) + nw//4); labels.append('Mid Pre-Production')
    if 'PRODUCTION' in pr:
        ms.append(ss); labels.append('Start of Production')
        ms.append(se); labels.append('End of Production')
    else:
        ms.append(nw//2); labels.append('Mid Production')
    if 'POST' in pr:
        ms.append(mid_post); labels.append('Mid Post')
    last = max(v[1] for v in pr.values())
    ms.append(last); labels.append('Final Delivery')

    combined = sorted(set(zip(ms, labels)), key=lambda x: x[0])
    ms = [x[0] for x in combined]; labels = [x[1] for x in combined]
    ranges = []
    for i in range(len(ms)-1):
        ranges.append((ms[i], ms[i+1]-1))
    ranges.append((ms[-1], nw-1))

    result = []
    for i, (s, e) in enumerate(ranges):
        total = sum(wk_totals[max(0,s):min(nw,e+1)])
        result.append({'label': f'Payment {i+1}', 'milestone': labels[i] if i < len(labels) else f'Payment {i+1}', 'start': s, 'end': e, 'amount': total})
    return result

# ── CASH FLOW EXCEL BUILDER ───────────────────────────────────────────────────


# ── DYNAMIC EXCEL BUILDER ─────────────────────────────────────────────────────

def build_excel(show_info, phases, sections):
    title   = show_info.get('showTitle','Untitled Show')
    network = show_info.get('network','')
    prod_co = show_info.get('prodCo','')

    def phase_start(p):
        d = parse_date(p['start'])
        return d if d else datetime(2099,1,1)
    phases = sorted(phases, key=phase_start)

    starts = [parse_date(p['start']) for p in phases if parse_date(p['start'])]
    ends   = [parse_date(p['end'])   for p in phases if parse_date(p['end'])]
    if not starts or not ends:
        raise ValueError('No valid phase dates found')

    wk1  = get_monday(min(starts))
    last = get_monday(max(ends))
    nw   = round((last - wk1).days / 7) + 1

    def wk_date(i):
        return wk1 + timedelta(weeks=i)

    pr = build_phase_ranges(phases, wk1, nw)

    # Build display_phases from the actual phases the user submitted
    # This makes the staircase dynamic — reflects whatever phases exist,
    # including custom ones (Run-Through, Development, etc.) and
    # excluding any the user deleted. Order follows chronological sort above.
    display_phases = []
    seen_names = set()
    for ph in phases:
        s_date = parse_date(ph['start'])
        e_date = parse_date(ph['end'])
        if not s_date or not e_date:
            continue
        name = ph['name'].upper().strip()
        if name in seen_names:
            continue
        seen_names.add(name)
        display_phases.append({
            'name':  ph['name'].upper(),
            'start': ph['start'],
            'end':   ph['end'],
        })

    line_data = []
    for s in sections:
        acct  = str(s.get('acct',''))
        label = str(s.get('label',''))
        raw   = s.get('total', 0)
        amt   = int(str(raw).replace(',','').replace('$','').split('.')[0]) if raw else 0
        stype = s.get('spread_type','full')
        wks   = spread(stype, amt, nw, pr)
        line_data.append((acct, label, amt, wks))

    # Guard: need at least one row or SUM formulas break
    if not line_data:
        line_data.append(('', 'No budget sections — add manually', 0, [0.0]*nw))

    wk_totals = [sum(ld[3][i] for ld in line_data) for i in range(nw)]
    payments  = compute_payments(wk_totals, nw, pr, phases)

    BOLD12 = Font(name='Arial', size=12, bold=True)
    REG12  = Font(name='Arial', size=12, bold=False)
    BOLD14 = Font(name='Arial', size=14, bold=True)
    GRAY   = PatternFill('solid', fgColor='C0C0C0')
    NOFILL = PatternFill(fill_type=None)
    CTR    = Alignment(horizontal='center', vertical='center')
    LFT    = Alignment(horizontal='left',   vertical='center')
    FMT    = '#,##0'

    def ap(cell, font=None, fill=None, align=None, fmt=None):
        if font:  cell.font          = font
        if fill:  cell.fill          = fill
        if align: cell.alignment     = align
        if fmt:   cell.number_format = fmt

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cash Flow'

    CA, CB, CC, CW1 = 1, 2, 3, 4
    CWL = CW1 + nw - 1
    CTT = CW1 + nw
    NP  = len(display_phases)
    NL  = len(line_data)
    NPM = len(payments)

    def cl(c): return get_column_letter(c)

    R_TITLE = 1; R_SUB = 2; R_WKN = 4; R_DAT = 5
    R_PH0   = 7
    R_HDR   = R_PH0 + NP + 1
    R_DS    = R_HDR + 2
    R_DE    = R_DS  + NL - 1
    R_TOT   = R_DE  + 2
    R_PS    = R_TOT + 3
    R_PE    = R_PS  + NPM - 1
    R_PSUM  = R_PE  + 2
    R_SHDR  = R_PSUM + 4
    R_SDAT  = R_SHDR + 2
    R_SEND  = R_SDAT + NPM - 1
    R_GRAND = R_SEND + 2

    run_date = datetime.now().strftime('%m/%d/%Y')
    ws.cell(R_TITLE, CA, 'CASH FLOW')
    ap(ws.cell(R_TITLE, CA), font=BOLD14, align=LFT)
    ws.cell(R_SUB, CA, ' | '.join(filter(None,[title,network,prod_co])) + f' | Generated {run_date}')
    ap(ws.cell(R_SUB, CA), font=BOLD12, align=LFT)
    ws.row_dimensions[R_TITLE].height = 22

    # No BUDGET TOTAL label in C4 — it appears in the header row below
    for i in range(nw):
        ap(ws.cell(R_WKN, CW1+i, f'WK {i+1}'), font=BOLD12, align=CTR)
    ap(ws.cell(R_WKN, CTT, 'TOTAL'), font=BOLD12, align=CTR)
    for i in range(nw):
        ap(ws.cell(R_DAT, CW1+i, wk_date(i).strftime('%m/%d/%y')), font=BOLD12, align=CTR)

    for pi, ph in enumerate(display_phases):
        r = R_PH0 + pi
        # Phase label in col C with colon, centered
        label = ph['name'].title() + ':'
        ap(ws.cell(r, CC, label), font=BOLD12, align=CTR)
        ws.row_dimensions[r].height = 16
        s_date = parse_date(ph['start']); e_date = parse_date(ph['end'])
        if s_date and e_date:
            ph_s = max(0, wk_idx(s_date, wk1))
            ph_e = min(nw-1, wk_idx(e_date, wk1))
            for i in range(nw):
                c = ws.cell(r, CW1+i)
                if ph_s <= i <= ph_e:
                    # Gray bar only — no text inside
                    ap(c, fill=GRAY)
                else:
                    ap(c, fill=NOFILL)

    for col, txt in [(CA,'ACCT#'),(CB,'DESCRIPTION'),(CC,'BUDGET TOTAL')]:
        ap(ws.cell(R_HDR, col, txt), font=BOLD12, align=CTR)
    for i in range(nw):
        ap(ws.cell(R_HDR, CW1+i, f'WK {i+1}'), font=BOLD12, align=CTR)
    ap(ws.cell(R_HDR, CTT, 'TOTAL'), font=BOLD12, align=CTR)
    ws.row_dimensions[R_HDR].height = 18

    wkf = cl(CW1); wkl = cl(CWL)
    for idx, (acct, label, budget, wks) in enumerate(line_data):
        r = R_DS + idx
        ap(ws.cell(r, CA, acct),   font=REG12, align=CTR)
        ap(ws.cell(r, CB, label),  font=REG12, align=LFT)
        ap(ws.cell(r, CC, budget), font=REG12, align=CTR, fmt=FMT)
        for i, v in enumerate(wks):
            rv = round(v)
            if rv != 0:
                ap(ws.cell(r, CW1+i, rv), font=REG12, align=CTR, fmt=FMT)
        c = ws.cell(r, CTT)
        c.value = f'=SUM({wkf}{r}:{wkl}{r})'
        ap(c, font=REG12, align=CTR, fmt=FMT)

    r = R_TOT; bc = cl(CC)
    ap(ws.cell(r, CB, 'TOTAL'), font=BOLD12, align=CTR)
    c = ws.cell(r, CC); c.value = f'=SUM({bc}{R_DS}:{bc}{R_DE})'
    ap(c, font=BOLD12, align=CTR, fmt=FMT)
    for i in range(nw):
        wc = cl(CW1+i); c = ws.cell(r, CW1+i)
        c.value = f'=SUM({wc}{R_DS}:{wc}{R_DE})'
        ap(c, font=REG12, align=CTR, fmt=FMT)
    c = ws.cell(r, CTT); c.value = f'=SUM({wkf}{r}:{wkl}{r})'
    ap(c, font=BOLD12, align=CTR, fmt=FMT)
    ws.row_dimensions[r].height = 18

    for pi, pmt in enumerate(payments):
        r = R_PS + pi
        ap(ws.cell(r, CB, pmt['label']), font=REG12, align=CTR)
        for i in range(nw):
            c = ws.cell(r, CW1+i)
            if pmt['start'] <= i <= pmt['end']: ap(c, fill=GRAY)
            else:                                ap(c, fill=NOFILL)
        fc = cl(CW1+pmt['start']); lc = cl(CW1+pmt['end'])
        tot_row = R_TOT
        formula = f'={lc}{tot_row}' if pmt['start'] == pmt['end'] else f'=+SUM({fc}{tot_row}:{lc}{tot_row})'
        c = ws.cell(r, CW1+pmt['end']); c.value = formula
        ap(c, font=REG12, align=CTR, fmt=FMT, fill=GRAY)
        c = ws.cell(r, CC);  c.value = f'={lc}{r}';                  ap(c, font=REG12, align=CTR, fmt=FMT)
        c = ws.cell(r, CTT); c.value = f'=SUM({wkf}{r}:{wkl}{r})'; ap(c, font=REG12, align=CTR, fmt=FMT)

    r = R_PSUM; bc = cl(CC); tc = cl(CTT)
    ap(ws.cell(r, CB, 'Payment Total'), font=REG12, align=CTR)
    c = ws.cell(r, CC);  c.value = f'=SUM({bc}{R_PS}:{bc}{R_PE})'; ap(c, font=REG12, align=CTR, fmt=FMT)
    c = ws.cell(r, CTT); c.value = f'=SUM({tc}{R_PS}:{tc}{R_PE})'; ap(c, font=REG12, align=CTR, fmt=FMT)

    ap(ws.cell(R_SHDR, CB, 'Payment Schedule'), font=BOLD12, align=CTR)
    ap(ws.cell(R_SHDR, CC, 'Milestone'),        font=BOLD12, align=CTR)
    ap(ws.cell(R_SHDR, CW1,   'Date'),   font=BOLD12, align=CTR)
    ap(ws.cell(R_SHDR, CW1+1, 'Amount'), font=BOLD12, align=CTR)
    for pi, pmt in enumerate(payments):
        r = R_SDAT + pi
        ap(ws.cell(r, CB, pmt['label']),     font=REG12, align=CTR)
        ap(ws.cell(r, CC, pmt['milestone']), font=REG12, align=CTR)
        ap(ws.cell(r, CW1, wk_date(min(pmt['start'], nw-1)).strftime('%m/%d/%y')), font=REG12, align=CTR)
        pr_row = R_PS + pi
        c = ws.cell(r, CW1+1); c.value = f'={cl(CC)}{pr_row}'; ap(c, font=REG12, align=CTR, fmt=FMT)

    r = R_GRAND; ac = cl(CW1+1)
    ap(ws.cell(r, CC, 'Total:'), font=BOLD12, align=CTR)
    c = ws.cell(r, CW1+1); c.value = f'=SUM({ac}{R_SDAT}:{ac}{R_SEND})'; ap(c, font=BOLD12, align=CTR, fmt=FMT)

    ws.column_dimensions[cl(CA)].width = 10
    ws.column_dimensions[cl(CB)].width = 32
    ws.column_dimensions[cl(CC)].width = 17
    for i in range(nw):
        ws.column_dimensions[cl(CW1+i)].width = 14
    ws.column_dimensions[cl(CTT)].width = 17
    ws.freeze_panes = ws.cell(R_DS, CW1)

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf, title

# ── AI PARSERS ────────────────────────────────────────────────────────────────
def _safe_json_parse(raw):
    clean = re.sub(r'^```[a-z]*\n?', '', raw.strip()).replace('```', '').strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        last_brace = clean.rfind('}')
        if last_brace > 0:
            trimmed = clean[:last_brace + 1]
            open_brackets = trimmed.count('[') - trimmed.count(']')
            open_braces   = trimmed.count('{') - trimmed.count('}')
            for _ in range(open_brackets): trimmed += ']'
            for _ in range(open_braces):   trimmed += '}'
            try:
                return json.loads(trimmed)
            except Exception:
                pass
        raise ValueError(f'JSON parse failed. Raw length: {len(raw)} chars. Preview: {raw[:300]}')


# ── COMBINED BUDGET + CALENDAR PARSER ────────────────────────────────────────

def parse_budget_and_calendar(budget_b64=None, calendar_b64=None):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=300.0)

    content_blocks = []
    doc_index = 1

    if budget_b64:
        content_blocks.append({
            'type': 'document',
            'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': budget_b64}
        })

    if calendar_b64:
        content_blocks.append({
            'type': 'document',
            'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': calendar_b64}
        })

    has_budget   = budget_b64   is not None
    has_calendar = calendar_b64 is not None

    budget_instructions = """
DOCUMENT 1 is a production budget.
Focus ONLY on the TOP SHEET — the summary page showing one total per department or account group.
Do NOT read the detail pages below the top sheet.

Extract ONLY the individual department/account section lines. These are the rows with a specific
account number and department name, like "100000 PRODUCERS" or "760000 POST PRODUCTION STAFF".

CRITICAL — DO NOT INCLUDE:
- Any row whose label contains the word "Total", "TOTAL", "Subtotal", or "Grand Total"
- Section group headers like "Above the Line", "Below the Line", "Post Production", "Other Expenses"
- Any summary or rollup row — only individual account lines with their own account numbers

Each valid section has its own unique account number. If two rows share or derive from the same
account numbers, only include the detail rows, not the summary.

For each valid section extract:
- acct: account number as string (e.g. "100000"), use "" if none visible
- label: section name exactly as shown (e.g. "PRODUCERS", "FIELD AUDIO", "ANIMALS")
- total: the dollar total for this section as an integer (0 if blank)

Also extract show metadata if visible: showTitle, network, prodCo, numEps (integer).
""" if has_budget else ""

    cal_doc_num = "2" if has_budget else "1"
    calendar_instructions = f"""
DOCUMENT {cal_doc_num} is a production calendar showing months across multiple pages.
Each page is a month. Phase bars appear inside date cells showing what is happening that week.

CRITICAL RULES:
1. Phase bars repeat week by week across pages — e.g. "Casting - Week 1", "Casting - Week 2" ... "Casting - Week 18"
   These are ALL the same phase. Collapse them: find the FIRST occurrence (earliest date) and the LAST occurrence (latest date).
2. Return ONE entry per canonical phase type, not one per week label.
3. The canonical phase types to look for (and their variants):
   - CASTING: any bar labeled "Casting", "Casting - Week N", "Talent Search", "Auditions"
   - PREP: any bar labeled "Prep", "Prep - Week N", "Pre-Production", "Office Open"
   - PRODUCTION: any bar labeled "Load In", "ESU", "Rehearse", "Shoot", "Shoot 101", "Dark", "Strike", "Hiatus"
   - POST: any bar labeled "Edit", "Edit - Week N", "Online", "Deliver", "RC", "FC", "LC", "Mix", "Color"
4. For each phase, the start date is the MONDAY of the first week that phase appears.
   The end date is the FRIDAY of the last week that phase appears.
5. IMPORTANT for PREP end date: Prep runs until the week BEFORE the first Production activity
   (Load In, ESU, Rehearse, or Shoot). Do not end Prep early just because Casting is still running.
   Prep and Casting overlap — that is correct and expected.
5. Phases can and do overlap — that is correct and expected.
6. Read ALL pages of the calendar before determining start and end dates.

Return one entry per canonical phase found:
- name: use the canonical name (CASTING, PREP, PRODUCTION, POST) not the week label
- start: YYYY-MM-DD
- end: YYYY-MM-DD
Sort by start date.
""" if has_calendar else ""

    spread_instructions = """
For each budget section, assign a spread_type based on the section name, account number, and the phases present in the calendar.

spread_type must be exactly one of:
casting | prep | production | post | full | week1 | prep_shoot | shoot_plus2 | shoot_plus4 | post_plus2 | graphics

Spread assignment rules (use these as your guide, applying judgment for unusual sections):
- Producers, showrunners, EPs, executive staff = full
- Rights, clearances, research = prep
- Office production staff, coordinators, accountants = prep_shoot
- Talent fees, host fees, participant fees = shoot_plus2
- Talent travel and expenses = prep
- Prize money = production
- Stunts, extras, stand-ins = production
- Casting directors, talent bookers = casting
- Directors, field producers = shoot_plus2
- Camera crews and equipment = shoot_plus2
- Audio crews and equipment = shoot_plus2
- Grip, electric, lighting = shoot_plus2
- Art department, scenic, set design = shoot_plus4
- Costume, wardrobe, makeup, hair = shoot_plus4
- Locations, craft service, catering = prep_shoot
- Transportation, vehicles = shoot_plus2
- Studio equipment, field equipment rentals = shoot_plus4
- Production travel (airfare, hotel, per diem) = shoot_plus4
- Set construction, studio operations = shoot_plus4
- General and administrative, office = full
- Production company fee, format fee, agency fee, legal fee = full
- Insurance, E&O, completion bond = week1
- Post production staff, editors, post supervisors = post_plus2
- Graphics, main titles, acquired footage = graphics
- Music, score = post
- Post finishing, online, color correction, audio mix = post
- Transcriptions, closed captioning = post
- Edit bays, post equipment, storage = post
- Master delivery, digital delivery = post
- Game show specific: gaming electronics, game props, prizes = production
- Animals, specialty performers = production
- Live audience costs = production
- Any section you cannot confidently categorize = full
""" if has_budget else ""

    prompt = budget_instructions + calendar_instructions + spread_instructions + """
Return ONLY valid JSON, no markdown, no explanation:
{
  "showTitle": "",
  "network": "",
  "prodCo": "",
  "numEps": 0,
  "sections": [
    {"acct": "100000", "label": "PRODUCERS", "total": 1500000, "spread_type": "full"}
  ],
  "phases": [
    {"name": "CASTING", "start": "2024-01-08", "end": "2024-03-01"}
  ]
}

Return empty arrays [] for sections or phases if that document was not provided.
Include ALL sections from the top sheet, even those with a zero total."""

    content_blocks.append({'type': 'text', 'text': prompt})

    response = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=8000,
        messages=[{'role': 'user', 'content': content_blocks}]
    )

    result = _safe_json_parse(response.content[0].text)

    # Snap all phase dates to Monday/Friday of their week
    def to_monday(date_str):
        if not date_str: return date_str
        try:
            d = parse_date(date_str)
            if d:
                return (d - timedelta(days=d.weekday())).strftime('%Y-%m-%d')
        except Exception:
            pass
        return date_str

    def to_friday(date_str):
        if not date_str: return date_str
        try:
            d = parse_date(date_str)
            if d:
                return (d + timedelta(days=(4 - d.weekday()) % 7)).strftime('%Y-%m-%d')
        except Exception:
            pass
        return date_str

    phases_out = result.get('phases', [])
    for ph in phases_out:
        ph['start'] = to_monday(ph.get('start',''))
        ph['end']   = to_friday(ph.get('end',''))

    # Business rule: Prep should run right up to Production start (no gap).
    # Find PREP and PRODUCTION phases and close the gap.
    def find_phase(name_fragment):
        for ph in phases_out:
            if name_fragment.upper() in ph.get('name','').upper():
                return ph
        return None

    prep = find_phase('PREP')
    prod = find_phase('PRODUCTION')
    if prep and prod and prod.get('start'):
        prod_start = parse_date(prod['start'])
        if prod_start:
            # Prep ends the Friday before Production Monday
            prep_end = prod_start - timedelta(days=3)  # Friday before Monday
            prep['end'] = prep_end.strftime('%Y-%m-%d')

    return result


def parse_budget_pdf(pdf_b64):
    """Legacy single-doc budget parse."""
    result = parse_budget_and_calendar(budget_b64=pdf_b64)
    out = {}
    for s in result.get('sections', []):
        if s.get('acct'):
            out[s['acct']] = s.get('total', 0)
    if result.get('showTitle'): out['_showTitle'] = result['showTitle']
    if result.get('network'):   out['_network']   = result['network']
    if result.get('prodCo'):    out['_prodCo']    = result['prodCo']
    if result.get('numEps'):    out['_numEps']    = result['numEps']
    return out


def parse_calendar_pdf(pdf_b64):
    """Legacy single-doc calendar parse."""
    result = parse_budget_and_calendar(calendar_b64=pdf_b64)
    phases = result.get('phases', [])
    def sort_key(p):
        d = parse_date(p.get('start',''))
        return d if d else datetime(2099,1,1)
    return sorted(phases, key=sort_key)

def parse_cost_report_pdf(pdf_b64):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=300.0)

    # Call 1: metadata + above-the-line (EA section, accounts 1xxxxx-2xxxxx)
    prompt_atl = """You are reading a TV/film production cost report PDF.

EXTRACT: Show metadata AND all Above the Line (EA section) line items only.
EA accounts start with: 100000, 120000, 140000, 160000, 170000, 180000, 200000.

For each line item return:
- acct: account number string (e.g. "100408")
- dept: department name (e.g. "PRODUCERS")
- description: line item label
- is_total: true if department subtotal row, false for detail lines
- actuals: actual spend integer (0 if blank)
- pos: purchase orders integer (0 if blank)
- efc: estimate at final cost integer — use EFC column, or Total Cost if no EFC (0 if blank)
- budget: approved budget integer (0 if blank)
- variance: variance integer if shown (null if not shown)

NOTE: overages may be shown as NEGATIVE variance values.

Return ONLY valid JSON, no markdown:
{
  "showTitle": "Show Name",
  "network": "NBC",
  "prodCo": "Production Co",
  "period": "6",
  "lines": [
    {"acct":"100408","dept":"PRODUCERS","description":"Co-Executive Producer","is_total":false,"actuals":349711,"pos":0,"efc":349711,"budget":370880,"variance":21169}
  ]
}
Use integers only. Include ALL EA lines including department subtotals."""

    # Call 2: below-the-line, other expenses, completion costs (EB/ED/EE)
    prompt_btl = """You are reading a TV/film production cost report PDF.

EXTRACT: All Below the Line (EB), Other Expenses (ED), and Completion Costs (EE) line items.
These include accounts starting with: 220000, 250000, 270000, 340000, 380000, 390000,
440000, 450000, 480000, 500000, 540000, 560000, 570000, 600000, 640000, 660000,
680000, 700000, 720000, 740000, 750000, 760000, 780000, 800000, 820000, 840000,
860000, 880000, and any other EB/ED/EE accounts. Also include Grand Total if present.

For each line item return:
- acct: account number string
- dept: department name
- description: line item label
- is_total: true if subtotal or grand total row
- actuals: actual spend integer (0 if blank)
- pos: purchase orders integer (0 if blank)
- efc: estimate at final cost integer (0 if blank)
- budget: approved budget integer (0 if blank)
- variance: variance integer if shown (null if not shown)

NOTE: overages may be shown as NEGATIVE variance values.

Return ONLY a valid JSON array, no markdown, no wrapper object:
[
  {"acct":"480432","dept":"GRIP/ELECTRIC","description":"Electrician","is_total":false,"actuals":143565,"pos":0,"efc":143565,"budget":211988,"variance":68423}
]
Use integers only. Include ALL EB/ED/EE lines including subtotals and grand total."""

    resp1    = client.messages.create(
        model='claude-opus-4-5', max_tokens=8000,
        messages=[{'role':'user','content':[
            {'type':'document','source':{'type':'base64','media_type':'application/pdf','data':pdf_b64}},
            {'type':'text','text':prompt_atl}
        ]}]
    )
    atl_data = _safe_json_parse(resp1.content[0].text)

    resp2    = client.messages.create(
        model='claude-opus-4-5', max_tokens=8000,
        messages=[{'role':'user','content':[
            {'type':'document','source':{'type':'base64','media_type':'application/pdf','data':pdf_b64}},
            {'type':'text','text':prompt_btl}
        ]}]
    )
    btl_raw  = _safe_json_parse(resp2.content[0].text)
    btl_lines = btl_raw if isinstance(btl_raw, list) else btl_raw.get('lines', [])

    return {
        'showTitle': atl_data.get('showTitle', ''),
        'network':   atl_data.get('network', ''),
        'prodCo':    atl_data.get('prodCo', ''),
        'period':    atl_data.get('period', ''),
        'lines':     atl_data.get('lines', []) + btl_lines
    }

# ── VARIANCE EXCEL BUILDER ────────────────────────────────────────────────────

def build_variance_excel(show_info, lines, threshold=10):
    title   = show_info.get('showTitle','Untitled Show')
    network = show_info.get('network','')
    prod_co = show_info.get('prodCo','')
    period  = show_info.get('period','')

    wb  = openpyxl.Workbook()

    # Colors
    RED        = PatternFill('solid', fgColor='FF4444')
    RED_ORANGE = PatternFill('solid', fgColor='FF7733')
    ORANGE     = PatternFill('solid', fgColor='FFB300')
    YELLOW     = PatternFill('solid', fgColor='FFE566')
    BLUE       = PatternFill('solid', fgColor='99CCFF')
    GRAY_HDR   = PatternFill('solid', fgColor='D0D0D0')
    GRAY_DEPT  = PatternFill('solid', fgColor='E8E8E8')
    NOFILL     = PatternFill(fill_type=None)

    BOLD14 = Font(name='Arial', size=13, bold=True)
    BOLD12 = Font(name='Arial', size=11, bold=True)
    BOLD11 = Font(name='Arial', size=11, bold=True)
    REG11  = Font(name='Arial', size=11)

    CTR = Alignment(horizontal='center', vertical='center')
    LFT = Alignment(horizontal='left',   vertical='center')
    RGT = Alignment(horizontal='right',  vertical='center')
    FMT = '#,##0'
    PCTFMT = '0.0%'

    def ap(cell, font=None, fill=None, align=None, fmt=None):
        if font:  cell.font          = font
        if fill:  cell.fill          = fill
        if align: cell.alignment     = align
        if fmt:   cell.number_format = fmt

    run_date   = datetime.now().strftime('%m/%d/%Y')
    period_str = f'Period {period}' if period else ''

    # ── SHEET 1: HOT SHEET ────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Hot Sheet'

    ws1.cell(1,1,'HOT SHEET — VARIANCE ANALYSIS')
    ap(ws1.cell(1,1), font=BOLD14, align=LFT)
    subtitle = ' | '.join(filter(None,[title,network,prod_co,period_str])) + f' | Generated {run_date}'
    ws1.cell(2,1, subtitle)
    ap(ws1.cell(2,1), font=BOLD11, align=LFT)

    key_data = [
        ('31%+ over budget',      RED),
        ('21-30% over budget',    RED_ORANGE),
        ('11-20% over budget',    ORANGE),
        ('1-10% over budget',     YELLOW),
        ('10%+ under budget',     BLUE),
        ('Within 10% (either)',   NOFILL),
    ]
    ws1.cell(1,7,'COLOR KEY'); ap(ws1.cell(1,7), font=BOLD11, align=LFT)
    for i,(label,fill) in enumerate(key_data):
        r = 2+i
        ap(ws1.cell(r,7), fill=fill)
        ws1.cell(r,8, label); ap(ws1.cell(r,8), font=REG11, align=LFT)

    HDR_ROW = 4
    headers = ['ACCT','DESCRIPTION','ACTUALS','POs','EFC','BUDGET','VARIANCE','VAR %','OVER/UNDER']
    for i,h in enumerate(headers):
        c = ws1.cell(HDR_ROW, i+1, h)
        ap(c, font=BOLD12, fill=GRAY_HDR, align=CTR)
    ws1.row_dimensions[HDR_ROW].height = 18

    widths = [10, 38, 16, 14, 16, 16, 16, 10, 14]
    for i,w in enumerate(widths):
        ws1.column_dimensions[get_column_letter(i+1)].width = w

    def color_for_row(budget, efc):
        if not budget or budget == 0:
            return NOFILL
        pct = (efc - budget) / abs(budget)
        if pct > 0.31:   return RED
        if pct > 0.21:   return RED_ORANGE
        if pct > 0.11:   return ORANGE
        if pct > 0.01:   return YELLOW
        if pct < -0.10:  return BLUE
        return NOFILL

    r = HDR_ROW + 1
    for line in lines:
        acct        = str(line.get('acct',''))
        dept        = str(line.get('dept',''))
        description = str(line.get('description',''))
        is_total    = line.get('is_total', False)
        actuals     = int(line.get('actuals') or 0)
        pos         = int(line.get('pos') or 0)
        efc         = int(line.get('efc') or 0)
        budget      = int(line.get('budget') or 0)
        variance    = line.get('variance')
        if variance is not None:
            variance = int(variance)

        font = BOLD11 if is_total else REG11
        fill = GRAY_DEPT if is_total else color_for_row(budget, efc)

        over_under = ''
        var_pct    = None
        if budget and budget != 0:
            var_pct    = (efc - budget) / abs(budget)
            over_under = 'OVER' if efc > budget else ('UNDER' if efc < budget else '')

        row_data = [acct, description if not is_total else f'{dept} TOTAL', actuals, pos, efc, budget,
                    variance if variance is not None else (efc - budget), var_pct, over_under]

        for ci, val in enumerate(row_data, 1):
            c = ws1.cell(r, ci, val)
            fmt_use = PCTFMT if ci == 8 else (FMT if ci in [3,4,5,6,7] else None)
            ap(c, font=font, fill=fill, align=(LFT if ci == 2 else CTR), fmt=fmt_use)

        r += 1

    # ── SHEET 2: NETWORK VARIANCE REPORT ─────────────────────────────────────
    ws2 = wb.create_sheet('Variance Report')

    ws2.cell(1,1,'VARIANCE REPORT')
    ap(ws2.cell(1,1), font=BOLD14, align=LFT)
    ws2.cell(2,1, subtitle)
    ap(ws2.cell(2,1), font=BOLD11, align=LFT)

    hdr2 = ['ACCT','DESCRIPTION','BUDGET','EFC','VARIANCE','VAR %','EXPLANATION']
    for i,h in enumerate(hdr2):
        c = ws2.cell(4, i+1, h)
        ap(c, font=BOLD12, fill=GRAY_HDR, align=CTR)
    ws2.row_dimensions[4].height = 18

    widths2 = [10, 38, 16, 16, 16, 10, 45]
    for i,w in enumerate(widths2):
        ws2.column_dimensions[get_column_letter(i+1)].width = w

    r2 = 5
    for line in lines:
        acct        = str(line.get('acct',''))
        dept        = str(line.get('dept',''))
        description = str(line.get('description',''))
        is_total    = line.get('is_total', False)
        efc         = int(line.get('efc') or 0)
        budget      = int(line.get('budget') or 0)
        variance    = line.get('variance')
        if variance is not None:
            variance = int(variance)
        else:
            variance = efc - budget

        font = BOLD11 if is_total else REG11
        fill = GRAY_DEPT if is_total else NOFILL

        var_pct = None
        if budget and budget != 0:
            var_pct = (efc - budget) / abs(budget)

        row_data2 = [acct, description if not is_total else f'{dept} TOTAL',
                     budget, efc, variance, var_pct, '']
        for ci, val in enumerate(row_data2, 1):
            c = ws2.cell(r2, ci, val)
            fmt_use = PCTFMT if ci == 6 else (FMT if ci in [3,4,5] else None)
            ap(c, font=font, fill=fill, align=(LFT if ci in [2,7] else CTR), fmt=fmt_use)

        r2 += 1

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf, title


# ── RISK & DILIGENCE SCANNER ──────────────────────────────────────────────────

RISK_CHUNK_SIZE = 80  # pages per AI call — safely under 100-page API limit

def _split_pdf_chunks(pdf_b64, chunk_size=RISK_CHUNK_SIZE):
    """Split a base64 PDF into chunks of chunk_size pages. Returns list of base64 strings."""
    raw = base64.b64decode(pdf_b64)
    reader = pypdf.PdfReader(io.BytesIO(raw))
    total = len(reader.pages)
    chunks = []
    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        writer = pypdf.PdfWriter()
        for i in range(start, end):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)
        chunks.append({
            'b64': base64.b64encode(buf.read()).decode(),
            'start_page': start + 1,
            'end_page': end,
            'total': total,
        })
    return chunks


def _call_risk_ai(client, pdf_b64, chunk_info, show_title=''):
    """Run one AI risk scan call on a PDF chunk."""
    page_context = f"(pages {chunk_info['start_page']}-{chunk_info['end_page']} of {chunk_info['total']})" if chunk_info['total'] > chunk_info['end_page'] else ""

    prompt = f"""You are a TV/film production risk analyst reading a production document {page_context}.
This could be a script, beat sheet, pitch deck, treatment, or any production document.

TASK 1 — Extract show metadata if visible (only return if found):
- showTitle, network, prodCo

TASK 2 — Flag every potential risk or diligence item you find.
Be thorough. It is better to over-flag than to miss something.

For each flag return:
- category: exactly one of:
    "IP & Legal" — brand names, logos, songs, shows, movies, real people, real companies,
                   real locations needing clearance, copyrighted formats, trademarked phrases
    "Physical Safety" — stunts, heights, water, fire, animals, extreme weather, vehicles,
                        weapons, crowds, confined spaces, food handling, pyrotechnics
    "Talent & Casting" — minors, medical conditions, nudity, intimacy, vulnerable participants
    "Location & Permits" — private property, government buildings, international shooting,
                           beaches, parks, airports, anything requiring permits
    "Broadcast Standards" — profanity, adult content, sensitive topics, gambling, alcohol, drugs
    "Insurance Triggers" — stunts, animals, watercraft, aircraft, pyrotechnics, anything
                           that would spike insurance or require special riders
    "Clearance & Research" — factual claims needing verification, historical events,
                             real crimes or legal cases, statistics or data cited

- severity: "High", "Medium", or "Low"
- location: page number or scene/section reference
- description: clear 1-2 sentence description of the specific risk
- quote: exact triggering words from the document (under 20 words)

Return ONLY valid JSON, no markdown:
{{
  "showTitle": "",
  "network": "",
  "prodCo": "",
  "flags": [
    {{
      "category": "IP & Legal",
      "severity": "High",
      "location": "Page 3",
      "description": "Description here.",
      "quote": "exact quote here"
    }}
  ]
}}

Include ALL flags. Sort by order of appearance."""

    response = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=8000,
        messages=[{'role': 'user', 'content': [
            {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': pdf_b64}},
            {'type': 'text', 'text': prompt}
        ]}]
    )
    return _safe_json_parse(response.content[0].text)


def parse_risk_document(pdf_b64=None, docx_b64=None):
    """Parse a script, beat sheet, pitch deck, or treatment for risks.
    Automatically chunks PDFs over 80 pages to stay within API limits."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=300.0)

    if not pdf_b64:
        return {'showTitle': '', 'network': '', 'prodCo': '', 'flags': []}

    # Check page count and chunk if needed
    raw = base64.b64decode(pdf_b64)
    reader = pypdf.PdfReader(io.BytesIO(raw))
    total_pages = len(reader.pages)

    if total_pages <= RISK_CHUNK_SIZE:
        # Single call — document fits within limit
        chunks = [{'b64': pdf_b64, 'start_page': 1, 'end_page': total_pages, 'total': total_pages}]
    else:
        # Split into chunks
        chunks = _split_pdf_chunks(pdf_b64)

    all_flags = []
    show_info = {'showTitle': '', 'network': '', 'prodCo': ''}

    for chunk in chunks:
        result = _call_risk_ai(client, chunk['b64'], chunk)
        # Take show info from first chunk that has it
        if not show_info['showTitle'] and result.get('showTitle'):
            show_info['showTitle'] = result['showTitle']
        if not show_info['network'] and result.get('network'):
            show_info['network'] = result['network']
        if not show_info['prodCo'] and result.get('prodCo'):
            show_info['prodCo'] = result['prodCo']
        all_flags.extend(result.get('flags', []))

    show_info['flags'] = all_flags
    return show_info

    prompt = """You are a TV/film production risk analyst reading a production document.
This could be a script, beat sheet, pitch deck, treatment, or any production document.

TASK 1 — Extract show metadata if visible:
- showTitle, network, prodCo

TASK 2 — Flag every potential risk or diligence item you find.
Be thorough. It is better to over-flag than to miss something.

For each flag return:
- category: one of these exact values:
    "IP & Legal" — brand names, logos, songs, shows, movies, real people, real companies,
                   real locations needing clearance, copyrighted formats, trademarked phrases
    "Physical Safety" — stunts, heights, water, fire, animals, extreme weather, vehicles,
                        weapons, crowds, confined spaces, food handling, pyrotechnics
    "Talent & Casting" — minors, medical conditions, nudity, intimacy, vulnerable participants
    "Location & Permits" — private property, government buildings, international shooting,
                           beaches, parks, airports, anything requiring permits
    "Broadcast Standards" — profanity, adult content, sensitive topics, gambling, alcohol, drugs
    "Insurance Triggers" — stunts, animals, watercraft, aircraft, pyrotechnics, anything
                           that would spike insurance or require special riders
    "Clearance & Research" — factual claims needing verification, historical events,
                             real crimes or legal cases, statistics or data cited

- severity: "High", "Medium", or "Low"
    High = legal exposure, serious safety risk, or show-stopper
    Medium = needs attention, likely requires action
    Low = worth noting, monitor, low urgency

- location: where in the document this appears — use page number if visible,
            otherwise scene name, section heading, or description like "Opening sequence"

- description: clear plain-English description of the specific risk, 1-2 sentences.
               Be specific — name the brand, person, location, or activity exactly as it appears.

- quote: the exact words or phrase from the document that triggered this flag (keep it short, under 20 words)

Also note: if the document appears to be a pitch deck with images, describe what you see visually.

Return ONLY valid JSON, no markdown:
{
  "showTitle": "Show Name",
  "network": "NBC",
  "prodCo": "Production Co",
  "flags": [
    {
      "category": "IP & Legal",
      "severity": "High",
      "location": "Page 3",
      "description": "The show format directly references and replicates The Amazing Race challenge structure.",
      "quote": "just like The Amazing Race, teams will race across..."
    }
  ]
}

Include ALL flags you find. Sort by order of appearance in the document."""

    content_blocks.append({'type': 'text', 'text': prompt})

    response = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=8000,
        messages=[{'role': 'user', 'content': content_blocks}]
    )

    return _safe_json_parse(response.content[0].text)


def build_risk_excel(show_info, flags):
    """Build a two-tab Excel: Document Order + By Category."""
    title   = show_info.get('showTitle', 'Untitled')
    network = show_info.get('network', '')
    prod_co = show_info.get('prodCo', '')

    wb = openpyxl.Workbook()

    # ── STYLES ────────────────────────────────────────────────────────────────
    BOLD12  = Font(name='Arial', size=11, bold=True)
    BOLD14  = Font(name='Arial', size=13, bold=True)
    REG11   = Font(name='Arial', size=11)
    REG10   = Font(name='Arial', size=10)
    CTR     = Alignment(horizontal='center', vertical='top', wrap_text=True)
    LFT     = Alignment(horizontal='left',   vertical='top', wrap_text=True)
    FMT_PCT = '0%'

    GRAY_HDR = PatternFill('solid', fgColor='D0D0D0')
    RED      = PatternFill('solid', fgColor='FF4444')
    ORANGE   = PatternFill('solid', fgColor='FFB300')
    YELLOW   = PatternFill('solid', fgColor='FFE566')
    NOFILL   = PatternFill(fill_type=None)

    SEV_FILL = {'High': RED, 'Medium': ORANGE, 'Low': YELLOW}
    SEV_FONT = {
        'High':   Font(name='Arial', size=10, bold=True, color='FFFFFF'),
        'Medium': Font(name='Arial', size=10, bold=True, color='000000'),
        'Low':    Font(name='Arial', size=10, bold=True, color='000000'),
    }

    CATEGORY_COLORS = {
        'IP & Legal':             'C5E0F5',
        'Physical Safety':        'FFD0D0',
        'Talent & Casting':       'D5F0D5',
        'Location & Permits':     'FFF0C0',
        'Broadcast Standards':    'EDD5F5',
        'Insurance Triggers':     'FFE0C0',
        'Clearance & Research':   'D5EEF5',
    }

    run_date = datetime.now().strftime('%m/%d/%Y')
    subtitle = ' | '.join(filter(None, [title, network, prod_co])) + f' | Generated {run_date}'

    HEADERS = ['#', 'CATEGORY', 'SEVERITY', 'LOCATION', 'DESCRIPTION', 'QUOTE']
    COL_WIDTHS = [5, 22, 12, 18, 55, 40]

    def write_sheet(ws, sheet_flags, sheet_title):
        # Title rows
        ws.cell(1, 1, sheet_title)
        ws.cell(1, 1).font = BOLD14
        ws.cell(2, 1, subtitle)
        ws.cell(2, 1).font = BOLD12

        # Headers
        for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            c = ws.cell(4, ci, h)
            c.font = BOLD12
            c.fill = GRAY_HDR
            c.alignment = CTR
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 20

        # Data rows
        for ri, flag in enumerate(sheet_flags, 5):
            cat      = flag.get('category', '')
            sev      = flag.get('severity', 'Low')
            loc      = flag.get('location', '')
            desc     = flag.get('description', '')
            quote    = flag.get('quote', '')
            cat_fill = PatternFill('solid', fgColor=CATEGORY_COLORS.get(cat, 'FFFFFF'))

            row_data = [ri - 4, cat, sev, loc, desc, quote]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(ri, ci, val)
                c.alignment = LFT if ci > 1 else CTR

                if ci == 2:  # Category
                    c.font = REG10
                    c.fill = cat_fill
                elif ci == 3:  # Severity
                    c.font = SEV_FONT.get(sev, REG10)
                    c.fill = SEV_FILL.get(sev, NOFILL)
                    c.alignment = CTR
                else:
                    c.font = REG10
                    c.fill = NOFILL

            ws.row_dimensions[ri].height = 40

        ws.freeze_panes = ws.cell(5, 1)

    # ── TAB 1: DOCUMENT ORDER ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Document Order'
    write_sheet(ws1, flags, 'RISK & DILIGENCE SCAN — Document Order')

    # ── TAB 2: BY CATEGORY ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('By Category')
    category_order = [
        'IP & Legal', 'Physical Safety', 'Insurance Triggers',
        'Talent & Casting', 'Location & Permits',
        'Broadcast Standards', 'Clearance & Research'
    ]
    # Sort by category order, then severity (High first), then original order
    sev_order = {'High': 0, 'Medium': 1, 'Low': 2}
    sorted_flags = sorted(
        flags,
        key=lambda f: (
            category_order.index(f.get('category','')) if f.get('category','') in category_order else 99,
            sev_order.get(f.get('severity','Low'), 3)
        )
    )
    write_sheet(ws2, sorted_flags, 'RISK & DILIGENCE SCAN — By Category')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf, title

# ── ROUTES ────────────────────────────────────────────────────────────────────



# ── RISK & DILIGENCE SCANNER ──────────────────────────────────────────────────

RISK_CHUNK_SIZE = 80  # pages per AI call — safely under 100-page API limit

def _split_pdf_chunks(pdf_b64, chunk_size=RISK_CHUNK_SIZE):
    """Split a base64 PDF into chunks of chunk_size pages. Returns list of base64 strings."""
    raw = base64.b64decode(pdf_b64)
    reader = pypdf.PdfReader(io.BytesIO(raw))
    total = len(reader.pages)
    chunks = []
    for start in range(0, total, chunk_size):
        end = min(start + chunk_size, total)
        writer = pypdf.PdfWriter()
        for i in range(start, end):
            writer.add_page(reader.pages[i])
        buf = io.BytesIO()
        writer.write(buf)
        buf.seek(0)
        chunks.append({
            'b64': base64.b64encode(buf.read()).decode(),
            'start_page': start + 1,
            'end_page': end,
            'total': total,
        })
    return chunks


def _call_risk_ai(client, pdf_b64, chunk_info, show_title=''):
    """Run one AI risk scan call on a PDF chunk."""
    page_context = f"(pages {chunk_info['start_page']}-{chunk_info['end_page']} of {chunk_info['total']})" if chunk_info['total'] > chunk_info['end_page'] else ""

    prompt = f"""You are a TV/film production risk analyst reading a production document {page_context}.
This could be a script, beat sheet, pitch deck, treatment, or any production document.

TASK 1 — Extract show metadata if visible (only return if found):
- showTitle, network, prodCo

TASK 2 — Flag every potential risk or diligence item you find.
Be thorough. It is better to over-flag than to miss something.

For each flag return:
- category: exactly one of:
    "IP & Legal" — brand names, logos, songs, shows, movies, real people, real companies,
                   real locations needing clearance, copyrighted formats, trademarked phrases
    "Physical Safety" — stunts, heights, water, fire, animals, extreme weather, vehicles,
                        weapons, crowds, confined spaces, food handling, pyrotechnics
    "Talent & Casting" — minors, medical conditions, nudity, intimacy, vulnerable participants
    "Location & Permits" — private property, government buildings, international shooting,
                           beaches, parks, airports, anything requiring permits
    "Broadcast Standards" — profanity, adult content, sensitive topics, gambling, alcohol, drugs
    "Insurance Triggers" — stunts, animals, watercraft, aircraft, pyrotechnics, anything
                           that would spike insurance or require special riders
    "Clearance & Research" — factual claims needing verification, historical events,
                             real crimes or legal cases, statistics or data cited

- severity: "High", "Medium", or "Low"
- location: page number or scene/section reference
- description: clear 1-2 sentence description of the specific risk
- quote: exact triggering words from the document (under 20 words)

Return ONLY valid JSON, no markdown:
{{
  "showTitle": "",
  "network": "",
  "prodCo": "",
  "flags": [
    {{
      "category": "IP & Legal",
      "severity": "High",
      "location": "Page 3",
      "description": "Description here.",
      "quote": "exact quote here"
    }}
  ]
}}

Include ALL flags. Sort by order of appearance."""

    response = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=8000,
        messages=[{'role': 'user', 'content': [
            {'type': 'document', 'source': {'type': 'base64', 'media_type': 'application/pdf', 'data': pdf_b64}},
            {'type': 'text', 'text': prompt}
        ]}]
    )
    return _safe_json_parse(response.content[0].text)


def parse_risk_document(pdf_b64=None, docx_b64=None):
    """Parse a script, beat sheet, pitch deck, or treatment for risks.
    Automatically chunks PDFs over 80 pages to stay within API limits."""
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY, timeout=300.0)

    if not pdf_b64:
        return {'showTitle': '', 'network': '', 'prodCo': '', 'flags': []}

    # Check page count and chunk if needed
    raw = base64.b64decode(pdf_b64)
    reader = pypdf.PdfReader(io.BytesIO(raw))
    total_pages = len(reader.pages)

    if total_pages <= RISK_CHUNK_SIZE:
        # Single call — document fits within limit
        chunks = [{'b64': pdf_b64, 'start_page': 1, 'end_page': total_pages, 'total': total_pages}]
    else:
        # Split into chunks
        chunks = _split_pdf_chunks(pdf_b64)

    all_flags = []
    show_info = {'showTitle': '', 'network': '', 'prodCo': ''}

    for chunk in chunks:
        result = _call_risk_ai(client, chunk['b64'], chunk)
        # Take show info from first chunk that has it
        if not show_info['showTitle'] and result.get('showTitle'):
            show_info['showTitle'] = result['showTitle']
        if not show_info['network'] and result.get('network'):
            show_info['network'] = result['network']
        if not show_info['prodCo'] and result.get('prodCo'):
            show_info['prodCo'] = result['prodCo']
        all_flags.extend(result.get('flags', []))

    show_info['flags'] = all_flags
    return show_info

    prompt = """You are a TV/film production risk analyst reading a production document.
This could be a script, beat sheet, pitch deck, treatment, or any production document.

TASK 1 — Extract show metadata if visible:
- showTitle, network, prodCo

TASK 2 — Flag every potential risk or diligence item you find.
Be thorough. It is better to over-flag than to miss something.

For each flag return:
- category: one of these exact values:
    "IP & Legal" — brand names, logos, songs, shows, movies, real people, real companies,
                   real locations needing clearance, copyrighted formats, trademarked phrases
    "Physical Safety" — stunts, heights, water, fire, animals, extreme weather, vehicles,
                        weapons, crowds, confined spaces, food handling, pyrotechnics
    "Talent & Casting" — minors, medical conditions, nudity, intimacy, vulnerable participants
    "Location & Permits" — private property, government buildings, international shooting,
                           beaches, parks, airports, anything requiring permits
    "Broadcast Standards" — profanity, adult content, sensitive topics, gambling, alcohol, drugs
    "Insurance Triggers" — stunts, animals, watercraft, aircraft, pyrotechnics, anything
                           that would spike insurance or require special riders
    "Clearance & Research" — factual claims needing verification, historical events,
                             real crimes or legal cases, statistics or data cited

- severity: "High", "Medium", or "Low"
    High = legal exposure, serious safety risk, or show-stopper
    Medium = needs attention, likely requires action
    Low = worth noting, monitor, low urgency

- location: where in the document this appears — use page number if visible,
            otherwise scene name, section heading, or description like "Opening sequence"

- description: clear plain-English description of the specific risk, 1-2 sentences.
               Be specific — name the brand, person, location, or activity exactly as it appears.

- quote: the exact words or phrase from the document that triggered this flag (keep it short, under 20 words)

Also note: if the document appears to be a pitch deck with images, describe what you see visually.

Return ONLY valid JSON, no markdown:
{
  "showTitle": "Show Name",
  "network": "NBC",
  "prodCo": "Production Co",
  "flags": [
    {
      "category": "IP & Legal",
      "severity": "High",
      "location": "Page 3",
      "description": "The show format directly references and replicates The Amazing Race challenge structure.",
      "quote": "just like The Amazing Race, teams will race across..."
    }
  ]
}

Include ALL flags you find. Sort by order of appearance in the document."""

    content_blocks.append({'type': 'text', 'text': prompt})

    response = client.messages.create(
        model='claude-opus-4-5',
        max_tokens=8000,
        messages=[{'role': 'user', 'content': content_blocks}]
    )

    return _safe_json_parse(response.content[0].text)


def build_risk_excel(show_info, flags):
    """Build a two-tab Excel: Document Order + By Category."""
    title   = show_info.get('showTitle', 'Untitled')
    network = show_info.get('network', '')
    prod_co = show_info.get('prodCo', '')

    wb = openpyxl.Workbook()

    # ── STYLES ────────────────────────────────────────────────────────────────
    BOLD12  = Font(name='Arial', size=11, bold=True)
    BOLD14  = Font(name='Arial', size=13, bold=True)
    REG11   = Font(name='Arial', size=11)
    REG10   = Font(name='Arial', size=10)
    CTR     = Alignment(horizontal='center', vertical='top', wrap_text=True)
    LFT     = Alignment(horizontal='left',   vertical='top', wrap_text=True)
    FMT_PCT = '0%'

    GRAY_HDR = PatternFill('solid', fgColor='D0D0D0')
    RED      = PatternFill('solid', fgColor='FF4444')
    ORANGE   = PatternFill('solid', fgColor='FFB300')
    YELLOW   = PatternFill('solid', fgColor='FFE566')
    NOFILL   = PatternFill(fill_type=None)

    SEV_FILL = {'High': RED, 'Medium': ORANGE, 'Low': YELLOW}
    SEV_FONT = {
        'High':   Font(name='Arial', size=10, bold=True, color='FFFFFF'),
        'Medium': Font(name='Arial', size=10, bold=True, color='000000'),
        'Low':    Font(name='Arial', size=10, bold=True, color='000000'),
    }

    CATEGORY_COLORS = {
        'IP & Legal':             'C5E0F5',
        'Physical Safety':        'FFD0D0',
        'Talent & Casting':       'D5F0D5',
        'Location & Permits':     'FFF0C0',
        'Broadcast Standards':    'EDD5F5',
        'Insurance Triggers':     'FFE0C0',
        'Clearance & Research':   'D5EEF5',
    }

    run_date = datetime.now().strftime('%m/%d/%Y')
    subtitle = ' | '.join(filter(None, [title, network, prod_co])) + f' | Generated {run_date}'

    HEADERS = ['#', 'CATEGORY', 'SEVERITY', 'LOCATION', 'DESCRIPTION', 'QUOTE']
    COL_WIDTHS = [5, 22, 12, 18, 55, 40]

    def write_sheet(ws, sheet_flags, sheet_title):
        # Title rows
        ws.cell(1, 1, sheet_title)
        ws.cell(1, 1).font = BOLD14
        ws.cell(2, 1, subtitle)
        ws.cell(2, 1).font = BOLD12

        # Headers
        for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
            c = ws.cell(4, ci, h)
            c.font = BOLD12
            c.fill = GRAY_HDR
            c.alignment = CTR
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[4].height = 20

        # Data rows
        for ri, flag in enumerate(sheet_flags, 5):
            cat      = flag.get('category', '')
            sev      = flag.get('severity', 'Low')
            loc      = flag.get('location', '')
            desc     = flag.get('description', '')
            quote    = flag.get('quote', '')
            cat_fill = PatternFill('solid', fgColor=CATEGORY_COLORS.get(cat, 'FFFFFF'))

            row_data = [ri - 4, cat, sev, loc, desc, quote]
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(ri, ci, val)
                c.alignment = LFT if ci > 1 else CTR

                if ci == 2:  # Category
                    c.font = REG10
                    c.fill = cat_fill
                elif ci == 3:  # Severity
                    c.font = SEV_FONT.get(sev, REG10)
                    c.fill = SEV_FILL.get(sev, NOFILL)
                    c.alignment = CTR
                else:
                    c.font = REG10
                    c.fill = NOFILL

            ws.row_dimensions[ri].height = 40

        ws.freeze_panes = ws.cell(5, 1)

    # ── TAB 1: DOCUMENT ORDER ─────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Document Order'
    write_sheet(ws1, flags, 'RISK & DILIGENCE SCAN — Document Order')

    # ── TAB 2: BY CATEGORY ────────────────────────────────────────────────────
    ws2 = wb.create_sheet('By Category')
    category_order = [
        'IP & Legal', 'Physical Safety', 'Insurance Triggers',
        'Talent & Casting', 'Location & Permits',
        'Broadcast Standards', 'Clearance & Research'
    ]
    # Sort by category order, then severity (High first), then original order
    sev_order = {'High': 0, 'Medium': 1, 'Low': 2}
    sorted_flags = sorted(
        flags,
        key=lambda f: (
            category_order.index(f.get('category','')) if f.get('category','') in category_order else 99,
            sev_order.get(f.get('severity','Low'), 3)
        )
    )
    write_sheet(ws2, sorted_flags, 'RISK & DILIGENCE SCAN — By Category')

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf, title

# ── ROUTES ────────────────────────────────────────────────────────────────────

@app.route('/health')
def health():
    return jsonify({'status':'ok','service':'StudioChief API'})

@app.route('/parse-budget', methods=['POST','OPTIONS'])
def parse_budget_route():
    if request.method == 'OPTIONS': return '', 200
    try:
        data    = request.get_json()
        pdf_b64 = data.get('pdf_base64', data.get('pdf_b64',''))
        if not pdf_b64: return jsonify({'error':'No PDF data provided'}), 400
        result = parse_budget_pdf(pdf_b64)
        return jsonify({'ok':True,'data':result})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/parse-calendar', methods=['POST','OPTIONS'])
def parse_calendar_route():
    if request.method == 'OPTIONS': return '', 200
    try:
        data    = request.get_json()
        pdf_b64 = data.get('pdf_base64', data.get('pdf_b64',''))
        if not pdf_b64: return jsonify({'error':'No PDF data provided'}), 400
        result = parse_calendar_pdf(pdf_b64)
        return jsonify({'ok':True,'phases':result})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/parse-both', methods=['POST','OPTIONS'])
def parse_both_route():
    """Combined budget + calendar — single AI call sees both PDFs together."""
    if request.method == 'OPTIONS': return '', 200
    try:
        data         = request.get_json()
        budget_b64   = data.get('budget_base64','') or None
        calendar_b64 = data.get('calendar_base64','') or None
        if not budget_b64 and not calendar_b64:
            return jsonify({'error':'No PDF data provided'}), 400
        result = parse_budget_and_calendar(budget_b64=budget_b64, calendar_b64=calendar_b64)
        return jsonify({'ok':True, 'data':result})
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/generate-cashflow', methods=['POST','OPTIONS'])
def generate_cashflow():
    if request.method == 'OPTIONS': return '', 200
    try:
        data      = request.get_json()
        show_info = data.get('showInfo',{})
        phases    = data.get('phases',[])
        sections  = data.get('sections', None)
        if sections is None:
            budget_vals = data.get('budgetValues',{})
            sections = [
                {'acct': k, 'label': k, 'total': v, 'spread_type': 'full'}
                for k, v in budget_vals.items() if v
            ]
        if not phases:
            return jsonify({'error':'No phases provided'}), 400
        buf, title = build_excel(show_info, phases, sections)
        safe_title = re.sub(r'[^\w\s\-]','',title).strip().replace(' ','_')
        filename   = f'{safe_title}_CashFlow_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

@app.route('/parse-risk', methods=['POST','OPTIONS'])
def parse_risk_route():
    if request.method == 'OPTIONS': return '', 200
    try:
        data     = request.get_json()
        pdf_b64  = data.get('pdf_base64','') or None
        docx_b64 = data.get('docx_base64','') or None
        if not pdf_b64 and not docx_b64:
            return jsonify({'error':'No document provided'}), 400

        result    = parse_risk_document(pdf_b64=pdf_b64, docx_b64=docx_b64)
        flags     = result.get('flags', [])
        show_info = {k: result.get(k,'') for k in ['showTitle','network','prodCo']}

        # Apply any manual show info overrides from request
        manual = data.get('show_info', {})
        for k in ['showTitle','network','prodCo']:
            if manual.get(k): show_info[k] = manual[k]

        buf, title = build_risk_excel(show_info, flags)
        safe_title = re.sub(r'[^\w\s\-]','',title).strip().replace(' ','_')
        filename   = f'{safe_title}_RiskScan_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(buf,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/parse-variance', methods=['POST','OPTIONS'])
def parse_variance_route():
    if request.method == 'OPTIONS': return '', 200
    try:
        data    = request.get_json()
        pdf_b64 = data.get('pdf_base64', data.get('pdf_b64',''))
        if not pdf_b64: return jsonify({'error':'No PDF data provided'}), 400
        cost_data = parse_cost_report_pdf(pdf_b64)
        lines     = cost_data.get('lines', [])
        show_info = {k: cost_data.get(k,'') for k in ['showTitle','network','prodCo','period']}
        buf, title = build_variance_excel(show_info, lines)
        safe_title = re.sub(r'[^\w\s\-]','',title).strip().replace(' ','_')
        filename   = f'{safe_title}_VarianceReport_{datetime.now().strftime("%Y%m%d")}.xlsx'
        return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        return jsonify({'error':str(e)}), 500

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 10000)), debug=False)
